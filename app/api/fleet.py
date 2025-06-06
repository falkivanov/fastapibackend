from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError
from app.database import get_db
from app.models.fleet import FleetVehicle
from app.models.schemas import FleetVehicleCreate, FleetVehicleOut, FleetVehicleUpdate
from typing import List
import openpyxl
import io
from fastapi.responses import StreamingResponse

router = APIRouter(prefix="/fleet", tags=["Fleet"])

@router.post("/", response_model=FleetVehicleOut)
def create_vehicle(vehicle: FleetVehicleCreate, db: Session = Depends(get_db)):
    db_vehicle = FleetVehicle(**vehicle.dict())
    db.add(db_vehicle)
    try:
        db.commit()
        db.refresh(db_vehicle)
        return db_vehicle
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Kennzeichen bereits vergeben.")

@router.get("/", response_model=List[FleetVehicleOut])
def list_active_vehicles(db: Session = Depends(get_db)):
    return db.query(FleetVehicle).filter(FleetVehicle.is_active == True).all()

@router.get("/{vehicle_id}", response_model=FleetVehicleOut)
def get_vehicle(vehicle_id: int, db: Session = Depends(get_db)):
    vehicle = db.query(FleetVehicle).filter(FleetVehicle.id == vehicle_id).first()
    if not vehicle:
        raise HTTPException(status_code=404, detail="Fahrzeug nicht gefunden")
    return vehicle

@router.put("/{vehicle_id}", response_model=FleetVehicleOut)
def update_vehicle(vehicle_id: int, updated: FleetVehicleUpdate, db: Session = Depends(get_db)):
    vehicle = db.query(FleetVehicle).filter(FleetVehicle.id == vehicle_id).first()
    if not vehicle:
        raise HTTPException(status_code=404, detail="Fahrzeug nicht gefunden")

    for field, value in updated.dict(exclude_unset=True).items():
        setattr(vehicle, field, value)

    db.commit()
    db.refresh(vehicle)
    return vehicle

@router.delete("/{vehicle_id}", status_code=204)
def delete_vehicle(vehicle_id: int, db: Session = Depends(get_db)):
    vehicle = db.query(FleetVehicle).filter(FleetVehicle.id == vehicle_id).first()
    if not vehicle:
        raise HTTPException(status_code=404, detail="Fahrzeug nicht gefunden")
    vehicle.is_active = False
    db.commit()
    return

@router.post("/upload_excel")
async def upload_vehicles_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    contents = await file.read()
    workbook = openpyxl.load_workbook(io.BytesIO(contents))
    sheet = workbook.active

    headers = [cell.value for cell in sheet[1]]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    inserted_count = 0

    for row in rows:
        vehicle_data = dict(zip(headers, row))

        new_vehicle = FleetVehicle(
            license_plate=vehicle_data.get("license_plate"),
            manufacturer=vehicle_data.get("manufacturer"),
            model=vehicle_data.get("model"),
            year=vehicle_data.get("year"),
            mileage=vehicle_data.get("mileage"),
            status=vehicle_data.get("status") or "active",
            is_active=True
        )
        db.add(new_vehicle)
        inserted_count += 1

    db.commit()
    return {"message": f"{inserted_count} Fahrzeuge erfolgreich importiert."}

@router.get("/export_template")
def export_vehicle_template():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["license_plate", "manufacturer", "model", "year", "mileage", "status"])

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=vehicle_template.xlsx"}
    )
