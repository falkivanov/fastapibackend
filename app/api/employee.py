from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse, JSONResponse
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError
from app.database import get_db
from app.models.employee import Employee
from app.models.schemas import EmployeeCreate, EmployeeOut, EmployeeUpdate
import openpyxl
import io
from datetime import datetime
from typing import List, Optional
from app.utils.date_utils import parse_date

# Konstanten
BATCH_SIZE = 1000
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

router = APIRouter(prefix="/employees", tags=["Employees"])

# Hilfsfunktion zum Parsen von Datumsstrings
def parse_date(date_str: Optional[str]) -> Optional[datetime.date]:
    if not date_str:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Unbekanntes Datumsformat: {date_str}")

# ============================================================
#                MITARBEITER ANLEGEN
# ============================================================

@router.post("/", response_model=EmployeeOut)
def create_employee(employee: EmployeeCreate, db: Session = Depends(get_db)):
    db_employee = Employee(**employee.dict())
    db.add(db_employee)
    try:
        db.commit()
        db.refresh(db_employee)
        return db_employee
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Transporter-ID '{employee.transporter_id}' existiert bereits.")

# ============================================================
#                ALLE AKTIVEN MITARBEITER LISTEN
# ============================================================

@router.get("/", response_model=List[EmployeeOut])
def list_employees(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db)
):
    return db.query(Employee).filter(Employee.is_active == True).offset(skip).limit(limit).all()

# ============================================================
#                EINZELNEN MITARBEITER ABFRAGEN
# ============================================================

@router.get("/{employee_id}", response_model=EmployeeOut)
def get_employee(employee_id: int, db: Session = Depends(get_db)):
    employee = db.query(Employee).filter(Employee.id == employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Mitarbeiter nicht gefunden")
    return employee

# ============================================================
#                MITARBEITER UPDATEN
# ============================================================

@router.put("/{employee_id}", response_model=EmployeeOut)
def update_employee(employee_id: int, updated_employee: EmployeeUpdate, db: Session = Depends(get_db)):
    employee = db.query(Employee).filter(Employee.id == employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Mitarbeiter nicht gefunden")

    for field, value in updated_employee.dict(exclude_unset=True).items():
        setattr(employee, field, value)

    try:
        db.commit()
        db.refresh(employee)
        return employee
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="Datenbankfehler beim Update")

# ============================================================
#                MITARBEITER LÖSCHEN
# ============================================================

@router.delete("/{employee_id}", status_code=204)
def delete_employee(employee_id: int, db: Session = Depends(get_db)):
    employee = db.query(Employee).filter(Employee.id == employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Mitarbeiter nicht gefunden")

    employee.is_active = False
    db.commit()
    return

# ============================================================
#                EXCEL UPLOAD: MITARBEITER IMPORT
# ============================================================

@router.post("/upload_excel")
async def upload_employees_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Nur Excel-Dateien sind erlaubt")
    
    if file.size > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="Datei zu groß")
    
    contents = await file.read()
    workbook = openpyxl.load_workbook(io.BytesIO(contents))
    sheet = workbook.active
    
    headers = [cell.value for cell in sheet[1]]
    required_fields = ["name", "transporter_id", "start_date"]
    
    # Validierung der Header
    for field in required_fields:
        if field not in headers:
            raise HTTPException(status_code=400, detail=f"Erforderliches Feld '{field}' fehlt in der Excel-Datei")
    
    employees = []
    inserted_count = 0
    error_rows = []
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            employee_data = dict(zip(headers, row))
            
            # Validierung der Pflichtfelder
            if not all(employee_data.get(field) for field in required_fields):
                error_rows.append(f"Zeile {row_idx}: Pflichtfelder fehlen")
                continue
                
            start_date = parse_date(employee_data.get("start_date"))
            end_date = parse_date(employee_data.get("end_date"))
            
            new_employee = Employee(
                name=employee_data.get("name"),
                email=employee_data.get("email"),
                phone=employee_data.get("phone"),
                telegram_username=employee_data.get("telegram_username"),
                transporter_id=employee_data.get("transporter_id"),
                mentor_first_name=employee_data.get("mentor_first_name"),
                mentor_last_name=employee_data.get("mentor_last_name"),
                start_date=start_date,
                end_date=end_date,
                days_per_week=employee_data.get("days_per_week"),
                is_flexible=employee_data.get("is_flexible"),
                prefers_six_days=employee_data.get("prefers_six_days"),
                vehicle=employee_data.get("vehicle"),
                address=employee_data.get("address"),
                federal_state=employee_data.get("federal_state"),
                is_active=True
            )
            employees.append(new_employee)
            inserted_count += 1
            
            if len(employees) >= BATCH_SIZE:
                db.bulk_save_objects(employees)
                db.commit()
                employees = []
                
        except Exception as e:
            error_rows.append(f"Zeile {row_idx}: {str(e)}")
    
    if employees:
        db.bulk_save_objects(employees)
        db.commit()
    
    return {
        "message": f"{inserted_count} Mitarbeiter erfolgreich importiert.",
        "errors": error_rows if error_rows else None
    }

# ============================================================
#                EXCEL EXPORT: ALLE MITARBEITER
# ============================================================

@router.get("/export")
def export_all_employees(
    skip: int = 0,
    limit: int = 1000,
    db: Session = Depends(get_db)
):
    employees = db.query(Employee).filter(Employee.is_active == True).offset(skip).limit(limit).all()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append([
        "name", "email", "phone", "telegram_username", "transporter_id",
        "mentor_first_name", "mentor_last_name", "start_date", "end_date",
        "days_per_week", "is_flexible", "prefers_six_days", "vehicle", "address",
        "federal_state"
    ])

    for emp in employees:
        sheet.append([
            emp.name, emp.email, emp.phone, emp.telegram_username, emp.transporter_id,
            emp.mentor_first_name, emp.mentor_last_name,
            emp.start_date.strftime("%d.%m.%Y") if emp.start_date else "",
            emp.end_date.strftime("%d.%m.%Y") if emp.end_date else "",
            emp.days_per_week, emp.is_flexible, emp.prefers_six_days,
            emp.vehicle, emp.address, emp.federal_state
        ])

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employees_export.xlsx"}
    )

# ============================================================
#                EXCEL EXPORT: LEERE VORLAGE
# ============================================================

@router.get("/export_template")
def export_employee_template():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append([
        "name", "email", "phone", "telegram_username", "transporter_id",
        "mentor_first_name", "mentor_last_name", "start_date", "end_date",
        "days_per_week", "is_flexible", "prefers_six_days", "vehicle", "address",
        "federal_state"
    ])

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employee_template.xlsx"}
    )
